import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def export_to_excel(data, output="assets.xlsx"):

    writer = pd.ExcelWriter(output, engine="openpyxl")

    severity_levels = ["critical", "high", "medium", "low"]
    sheet_created = False

    for sev in severity_levels:

        rows = [d for d in data if d["severity"] == sev]

        if not rows:
            continue

        formatted_rows = []

        # ---------------------------
        # GROUPING (vuln → vlan → assets)
        # ---------------------------
        vuln_group = {}

        for r in rows:
            vuln = r["vulnerability"]
            vlan = r["report"]

            vuln_group.setdefault(vuln, {}).setdefault(vlan, []).append({
                "asset": r["asset"],
                "status": r.get("status", "")
            })

        # ---------------------------
        # BUILD ROWS (3 assets per row)
        # ---------------------------
        for vuln, vlan_dict in vuln_group.items():
            for vlan, assets in vlan_dict.items():
                assets = sorted(assets, key=lambda x: x["asset"])
                for i in range(0, len(assets), 3):
                    chunk = assets[i:i+3]

                    formatted_rows.append({
                        "Vulnerability": vuln,
                        "VLAN": vlan,
                        "Asset 1": chunk[0]["asset"] if len(chunk) > 0 else "",
                        "Asset 2": chunk[1]["asset"] if len(chunk) > 1 else "",
                        "Asset 3": chunk[2]["asset"] if len(chunk) > 2 else "",
                        "Status": "Plugin Available" if any(c["status"] for c in chunk) else ""
                    })
        formatted_rows = sorted(
            formatted_rows,
            key=lambda x: (x["Vulnerability"], x["VLAN"])
        )
        df = pd.DataFrame(formatted_rows)
        df.to_excel(writer, sheet_name=sev.capitalize(), index=False)
        sheet_created = True

    if not sheet_created:
        pd.DataFrame([{"Message": "No vulnerabilities found"}]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    writer.close()

    # ---------------------------
    # 🔥 STYLING + MERGING (FIXED)
    # ---------------------------
    wb = load_workbook(output)

    fill1 = PatternFill(start_color="DCEAF7", end_color="DCEAF7", fill_type="solid")
    fill2 = PatternFill(start_color="B7D3F2", end_color="B7D3F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        current_vuln = None
        vuln_start = 2
        current_status = None
        status_start = 2
        current_vlan = None
        vlan_start = 2

        toggle = False

        for row in range(2, ws.max_row + 2):

            vuln = ws.cell(row=row, column=1).value
            vlan = ws.cell(row=row, column=2).value

            # -----------------------
            # VULNERABILITY MERGE (FIXED)
            # -----------------------
            if vuln != current_vuln:

                if current_vuln is not None:
                    ws.merge_cells(start_row=vuln_start, start_column=1,
                                   end_row=row-1, end_column=1)
                    ws.merge_cells(start_row=status_start, start_column=6,
                       end_row=row-1, end_column=6)

                current_vuln = vuln
                vuln_start = row
                toggle = not toggle
                status_start = row
                current_vlan = None  # 🔥 RESET VLAN

            # -----------------------
            # VLAN MERGE (FIXED)
            # -----------------------
            if vlan != current_vlan:

                if current_vlan is not None:
                    ws.merge_cells(start_row=vlan_start, start_column=2,
                                   end_row=row-1, end_column=2)

                current_vlan = vlan
                vlan_start = row

            # -----------------------
            # STRIP COLOR + BORDER
            # -----------------------
            fill = fill1 if toggle else fill2

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border

        # -----------------------
        # FINAL MERGES
        # -----------------------
        ws.merge_cells(start_row=vuln_start, start_column=1,
                       end_row=ws.max_row, end_column=1)

        ws.merge_cells(start_row=vlan_start, start_column=2,
                       end_row=ws.max_row, end_column=2)

        # -----------------------
        # HEADER STYLE
        # -----------------------
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.border = thin_border

        # -----------------------
        # COLUMN WIDTH
        # -----------------------
        widths = [45, 28, 28, 28, 28, 15]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

        # ALIGNMENT
        for cell in ws["A"]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for cell in ws["B"]:
            cell.alignment = Alignment(vertical="center")

    wb.save(output)

    print("[+] FINAL PERFECT Excel generated 🚀")