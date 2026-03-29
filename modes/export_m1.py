import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def export_to_excel(data, output="assets.xlsx"):
    writer = pd.ExcelWriter(output, engine="openpyxl")

    severity_levels = ["critical", "high", "medium", "low"]
    sheet_created = False
    
    # We will store the structure of each sheet to know exactly where to merge later
    sheet_merge_instructions = {}

    for sev in severity_levels:
        rows = [d for d in data if d["severity"] == sev]

        if not rows:
            continue

        formatted_rows = []
        vuln_group = {}

        # ---------------------------
        # GROUPING (vuln → vlan → assets)
        # ---------------------------
        for r in rows:
            vuln = r["vulnerability"]
            vlan = r["report"]

            vuln_group.setdefault(vuln, {}).setdefault(vlan, []).append({
                "asset": r["asset"],
                "status": r.get("status", "")
            })

        # ---------------------------
        # BUILD ROWS & TRACK SPANS
        # ---------------------------
        # openpyxl is 1-based, and headers are row 1, so data starts at row 2
        current_row = 2 
        vuln_merges = []
        vlan_merges = []

        # Sort vulnerabilities to ensure a predictable, sorted output
        for vuln in sorted(vuln_group.keys()):
            vuln_start_row = current_row
            
            for vlan in sorted(vuln_group[vuln].keys()):
                assets = sorted(vuln_group[vuln][vlan], key=lambda x: x["asset"])
                vlan_start_row = current_row
                
                for i in range(0, len(assets), 3):
                    chunk = assets[i:i+3]

                    formatted_rows.append({
                        "Vulnerability": vuln,
                        "VLAN": vlan,
                        "Asset 1": chunk[0]["asset"] if len(chunk) > 0 else "",
                        "Asset 2": chunk[1]["asset"] if len(chunk) > 1 else "",
                        "Asset 3": chunk[2]["asset"] if len(chunk) > 2 else "",
                        "Status": "Plugin Available" if any(c["status"] for c in chunk) else ""
                    })
                    current_row += 1
                
                # Save the start and end row for this specific VLAN
                if current_row - 1 > vlan_start_row:
                    vlan_merges.append((vlan_start_row, current_row - 1))
            
            # Save the start and end row for this specific Vulnerability
            if current_row - 1 > vuln_start_row:
                vuln_merges.append((vuln_start_row, current_row - 1))

        df = pd.DataFrame(formatted_rows)
        df.to_excel(writer, sheet_name=sev.capitalize(), index=False)
        
        # Store instructions for the styling block
        sheet_merge_instructions[sev.capitalize()] = {
            "vuln_merges": vuln_merges,
            "vlan_merges": vlan_merges
        }
        sheet_created = True

    if not sheet_created:
        pd.DataFrame([{"Message": "No vulnerabilities found"}]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    writer.close()

    # ---------------------------
    # 🔥 STYLING + MERGING (SOLVED)
    # ---------------------------
    wb = load_workbook(output)

    fill1 = PatternFill(start_color="DCEAF7", end_color="DCEAF7", fill_type="solid")
    fill2 = PatternFill(start_color="B7D3F2", end_color="B7D3F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in wb.sheetnames:
        if sheet == "Summary":
            continue
            
        ws = wb[sheet]
        instructions = sheet_merge_instructions.get(sheet, {"vuln_merges": [], "vlan_merges": []})

        # Apply basic cell colors and borders first
        toggle = False
        current_vuln_val = None
        
        for row in range(2, ws.max_row + 1):
            vuln_val = ws.cell(row=row, column=1).value
            if vuln_val != current_vuln_val:
                toggle = not toggle
                current_vuln_val = vuln_val
                
            fill = fill1 if toggle else fill2
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border

        # Apply VLAN merges (Column 2)
        for start, end in instructions["vlan_merges"]:
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)

        # Apply Vulnerability (Col 1) and Status (Col 6) merges
        for start, end in instructions["vuln_merges"]:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.merge_cells(start_row=start, start_column=6, end_row=end, end_column=6)

        # -----------------------
        # HEADER STYLE
        # -----------------------
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.border = thin_border

        # -----------------------
        # COLUMN WIDTH & ALIGNMENT
        # -----------------------
        widths = [45, 28, 28, 28, 28, 15]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

        for row in range(1, ws.max_row + 1):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                # Keep top-left alignment for merged cells to look good
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    wb.save(output)
    print("[+] Excel generated 🚀")