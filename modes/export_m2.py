import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import socket

# Helper function to sort IP strings numerically instead of alphabetically
def ip_sort_key(ip_string):
    try:
        # Converts "192.168.1.1" to a packed byte string for accurate sorting
        return socket.inet_aton(ip_string)
    except socket.error:
        # Fallback for hostnames or string data
        return ip_string


def export_to_excel(data, output="vlan_assets.xlsx"):

    writer = pd.ExcelWriter(output, engine="openpyxl")
    
    # -----------------------------------------------
    # 1. PARSE & GROUP BY VLAN -> IP -> PORTS
    # -----------------------------------------------
    structured_data = {}
    
    for r in data:
        vlan = r.get("report", "Unknown_VLAN")
        full_asset = r.get("asset", "")
        
        if not full_asset:
            continue
            
        parts = full_asset.split()
        ip = parts[0]
        
        port = ""
        if len(parts) > 1:
            port = parts[1].replace("(", "").replace(")", "").replace("tcp/", "").replace("udp/", "")
            
        if not port:
            port = "0" 
            
        structured_data.setdefault(vlan, {}).setdefault(ip, set()).add(port)

    sheet_created = False
    sheet_merge_instructions = {}
    vlan_sheet_mapping = {}

    # -----------------------------------------------
    # 2. BUILD SHEETS PER VLAN
    # -----------------------------------------------
    for vlan in sorted(structured_data.keys()):
        ip_dict = structured_data[vlan]
        formatted_rows = []
        current_row = 2 
        
        ip_merges = []
        vlan_start_row = 2
        
        sorted_ips = sorted(ip_dict.keys(), key=ip_sort_key)
        
        for ip in sorted_ips:
            sorted_ports = sorted(list(ip_dict[ip]), key=lambda x: int(x) if x.isdigit() else 99999)
            ip_start_row = current_row
            
            for port in sorted_ports:
                formatted_rows.append({
                    "vlan": vlan,
                    "ip": ip,
                    # 🔥 CHANGED: If port is "0" (unknown), leave it blank
                    "port": port if port != "0" else ""
                })
                current_row += 1
                
            if current_row - 1 > ip_start_row:
                ip_merges.append((ip_start_row, current_row - 1))
                
        if not formatted_rows:
            continue
            
        sheet_name = str(vlan).replace(":", "_").replace("/", "_")[:31]
        vlan_sheet_mapping[vlan] = sheet_name
        
        df = pd.DataFrame(formatted_rows)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        sheet_merge_instructions[sheet_name] = {
            "vlan_span": (vlan_start_row, current_row - 1),
            "ip_merges": ip_merges
        }
        sheet_created = True

    # -----------------------------------------------
    # 3. BUILD THE INDEX PAGE
    # -----------------------------------------------
    if sheet_created:
        index_rows = []
        for vlan in sorted(vlan_sheet_mapping.keys()):
            sheet_name = vlan_sheet_mapping[vlan]
            formula = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{vlan}")'
            index_rows.append({"VLAN Index (Click to Navigate)": formula})
            
        index_df = pd.DataFrame(index_rows)
        index_df.to_excel(writer, sheet_name="Index", index=False)
    else:
        pd.DataFrame([{"vlan": "None", "ip": "None", "port": "None"}]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    writer.close()

    # -----------------------------------------------
    # 4. STYLING & TARGETED MERGES
    # -----------------------------------------------
    wb = load_workbook(output)

    if "Index" in wb.sheetnames:
        sheet_list = wb.sheetnames
        sheet_list.remove("Index")
        sheet_list.insert(0, "Index")
        wb._sheets = [wb[s] for s in sheet_list]

    fill1 = PatternFill(start_color="F2F7FC", fill_type="solid")
    fill2 = PatternFill(start_color="DCEAF7", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        if sheet == "Index":
            ws.column_dimensions['A'].width = 50
            header = ws.cell(row=1, column=1)
            header.font = Font(bold=True, size=12)
            header.alignment = Alignment(horizontal="center", vertical="center")
            header.border = thin_border
            header.fill = PatternFill(start_color="B7D3F2", fill_type="solid")
            
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=1)
                cell.font = Font(color="0000FF", underline="single")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill1 if row % 2 == 0 else fill2
            continue
            
        if sheet == "Summary":
            continue
            
        instructions = sheet_merge_instructions.get(sheet, {"vlan_span": (2, 2), "ip_merges": []})

        toggle = False
        current_ip_val = None
        
        for row in range(2, ws.max_row + 1):
            ip_val = ws.cell(row=row, column=2).value
            if ip_val != current_ip_val:
                toggle = not toggle
                current_ip_val = ip_val
                
            fill = fill1 if toggle else fill2
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        v_start, v_end = instructions["vlan_span"]
        if v_end > v_start:
            ws.merge_cells(start_row=v_start, start_column=1, end_row=v_end, end_column=1)

        for i_start, i_end in instructions["ip_merges"]:
            ws.merge_cells(start_row=i_start, start_column=2, end_row=i_end, end_column=2)

        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = [35, 25, 15]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

    try:
        wb.save(output)
        print(f"[+] Cleaned layout with Index generated successfully in '{output}' 🚀")
    except PermissionError:
        print(f"\n❌ Error: Could not save '{output}'. Close the file in Excel and try again.\n")